import copy
import csv
import hashlib
import json
import logging
import mimetypes
import operator
import os
import sys
import time
import traceback
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO

import numpy as np
import openpyxl
import pandas as pd
import requests
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.core.management import call_command
from django.db import transaction
from django.db.models.fields.files import ImageFieldFile
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.http import (Http404, HttpResponse, HttpResponseBadRequest,
                         JsonResponse)
from django.shortcuts import HttpResponseRedirect, redirect, render
from django.utils import timezone
from django.views.generic.base import TemplateView, View
from drf_spectacular.utils import (OpenApiCallback, OpenApiExample,
                                   OpenApiParameter, OpenApiResponse,
                                   Serializer, extend_schema,
                                   inline_serializer)
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from rest_framework import authentication, generics, permissions, serializers
from rest_framework.authentication import SessionAuthentication
from rest_framework.authtoken.models import Token
from rest_framework.exceptions import ParseError
from rest_framework.response import Response
from rest_framework.views import APIView

from common.util import (aes_decrypt, aes_encrypt, check_role,
                         excel_file_write_sheets, get_sub_domain, pdata_handle)
from diablo.models import BkeyRegnoList, LkeyRegnoList
from i_search.forms import UploadFileForm
from i_search.management.commands import handle_datas, owner_handle
from i_search.models import Abiu, Babaco, Citron, Damson, History, TIRecord
from i_search.serializers import (AddRemarkSerializer,
                                  DownloadPhoneNumberSerializer,
                                  SaveRecordSerializer, TirecordSerializer)
from users.models import Company, CompanyUserMapping, DownloadHistory, User

logger = logging.getLogger(__name__)
is_test = False
#! 上code請遮
# is_test = True

file_path = "output_file"

def download(request, filename):
    filepath = os.path.join(file_path, filename)
    # print(filepath)
    # print(os.path.exists(filepath))
    if os.path.exists(filepath):
        if request.user.is_authenticated:
            with open(filepath, 'rb') as fh:
                mime_type, _ = mimetypes.guess_type(filepath)
                response = HttpResponse(fh.read(), content_type=mime_type)
                response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(filepath)
                return response
    raise Http404

def company_newest_history(ptoken, user_id):
    o_uid = user_id
    company_id = CompanyUserMapping.objects.filter(user_id=user_id, is_valid=1)[0].company_id
    user_datas = CompanyUserMapping.objects.filter(company_id=company_id, is_valid=1).values('user_id')
    user_list = []
    for i in user_datas:
        user_id = i['user_id']
        if not user_id in user_list:
            user_list.append(user_id)
    ti_list = [i['id'] for i in TIRecord.objects.filter(ptoken=ptoken, user_id__in=user_list).values('id')]
    memo = History.objects.filter(ti_record_id__in=ti_list).order_by('-id')[0].memo

    user_list.remove(o_uid)
    ti_update_unread = TIRecord.objects.filter(ptoken=ptoken, user_id__in=user_list)
    return memo, ti_update_unread

class CsrfExemptSessionAuthentication(SessionAuthentication):
    def enforce_csrf(self, request):
        return None

class CustomJsonEncoder(json.JSONEncoder):
    def default(self, obj): # pylint: disable=E0202
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, datetime):
            return obj.isoformat()
        elif isinstance(obj, date):
            return obj.strftime("%Y-%m-%d")
        elif isinstance(obj, ImageFieldFile):
            if obj:
                return obj.path
            return None
        return json.JSONEncoder.default(self, obj)

class NormalTemplateView(TemplateView):
    TAG = '[NormalTemplateView]'

    def get_context_data(self, **kwargs):
        context = super(NormalTemplateView, self).get_context_data(**kwargs)
        if self.request.GET.get('next'):
            logger.debug('I see next!!!')
            context['next'] = self.request.GET.get('next')
        if self.request.GET.get('display_main'):
            logger.debug('I see display_main!!!')
            context['display_main'] = self.request.GET.get('display_main')
        debug = False
        if not self.request.user.is_anonymous:
            context['first_name'] = self.request.user.first_name if len(self.request.user.first_name) > 0 else None
            context['user_name'] = self.request.user.username
            if settings.DEBUG == True:
                debug = True
            user_token = Token.objects.get(user=self.request.user).key
            context = {"user_token": user_token}
        context['debug'] = debug
        return context

class IndexView(NormalTemplateView):
    TAG = '[IndexView]'
    template_name = 'index.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if not self.request.user.is_anonymous:
            return redirect('/i_search/telem/')
        return render(request, self.template_name, context=context)

class TelemView(NormalTemplateView):
    TAG = '[TelemView]'
    template_name = 'telem.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        context['timestamp'] = datetime.now().timestamp()
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context)

class GetAreaListView(APIView):
    authentication_classes = [authentication.SessionAuthentication]
    permission_classes = [permissions.AllowAny]
    @extend_schema(
        summary='取行政區',
        description='取行政區',
        request=None,
        responses={
            200: OpenApiResponse(description='ok'),
            401: OpenApiResponse(description='身分認證失敗'),
        },
    )
    def get(self, request):
        res = {}
        city_code = request.GET.get('city')
        url = f'https://lbor.wsos.com.tw/common/car/get_area/?city={city_code}'
        result = requests.get(url)
        res = json.loads(result.text)
        return Response(res)

class AccountAddView(NormalTemplateView):
    TAG = '[AccountAddView]'
    template_name = 'account_add.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context)

class AccountManageView(NormalTemplateView):
    TAG = '[AccountManageView]'
    template_name = 'account_manage.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context)

class AccountEditView(NormalTemplateView):
    TAG = '[AccountEditView]'
    template_name = 'account_edit.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context)

class MemberAclistView(NormalTemplateView):
    TAG = '[MemberAclistView]'
    template_name = 'member_aclist.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context)
class ListDownloadView(NormalTemplateView):
    TAG = '[ListDownloadView]'
    template_name = 'list_download.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context) 
class AccountDownloadView(NormalTemplateView):
    TAG = '[AccountDownloadView]'
    template_name = 'account_download.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context)                

class MemberNewacView(NormalTemplateView):
    TAG = '[MemberNewacView]'
    template_name = 'member_newac.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context)

class MemberEditacView(NormalTemplateView):
    TAG = '[MemberEditacView]'
    template_name = 'member_editac.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(*args, **kwargs)
        if self.request.user.is_anonymous:
            return redirect('/')
        return render(request, self.template_name, context=context)

class LoginView(View):
    # authentication_classes = (CsrfExemptSessionAuthentication, )
    def post(self, request):
        result = {"status": "NG", "msg": 'not login'}
        try:
            sub_domain = get_sub_domain(request)
            account = request.POST.get('login')
            passwd = request.POST.get('password')
            remember = request.POST.get('remember')
            user = authenticate(username=account, password=passwd)
            if not user:
                account = ''.join([sub_domain, '.', account])
                user = authenticate(username=account, password=passwd)
            # check_admin = CompanyUserMapping.objects.filter(user_id=user.id, is_valid=1, is_admin=1)
            # if account and not account in ['dennis', 'coder', 'kevin', 'sunny', 'richer', 'ian', 'quinn'] and not check_admin:
            #     account = ''.join([sub_domain, '.', account])
            if user is not None:
                login(request, user)
                if remember !='true':
                    request.session.set_expiry(0)
                    request.session.modified = True
                result = {"status": "OK", "msg": 'login'}
                logger.debug(result)
                return HttpResponse(json.dumps(result, ensure_ascii=False), content_type="application/json; charset=utf-8")
        except Exception as error:
            result['error'] = "{}".format(error)
            logger.debug(result)
        return HttpResponseBadRequest(json.dumps(result, ensure_ascii=False), content_type="application/json; charset=utf-8")

class LogoutView(View):

    def get(self, request):
        logout(request)
        result = {'status': 'OK', 'msg': 'logout'}
        return HttpResponse(json.dumps(result, ensure_ascii=False), content_type="application/json; charset=utf-8")

class GetRecordInfo(APIView):
    TAG = "[GetRecordInfo]"

    authentication_classes = (CsrfExemptSessionAuthentication, )

    def none_data(self, phone=None):
        data_list = [{
                'user_info': {
                            'uid': None,
                            'name': None,
                            'bday': None,
                            'gender': None,
                            'cell_phone': phone,
                            'phone': None,
                            'addr': None,
                            'address_current': None,
                            'uid_tag': None
                            },
                'land': [{
                        'city_name': None,
                        'area_name': None,
                        'region_name': None,
                        'lno': None,
                        'regno': None,
                        'reg_date_str': None,
                        'reg_reason_str': None,
                        'right_str': None,
                        'shared_size': None,
                        'present_value': None,
                        'owner_data': None
                        }],
                'build': [{
                        'city': None,
                        'district': None,
                        'segment': None,
                        'buildName': None,
                        'houseNumber': None,
                        'otherAuthority': None,
                        'regno': None,
                        'reg_date_str': None,
                        'reg_reason_str': None,
                        'totalArea': None,
                        'masterArea': None,
                        'slaveArea': None,
                        'publicArea': None,
                        'carArea': None,
                        'floorNum': None,
                        'ownershipMark': None,
                        'buildNum': None
                        }],
                'Property': [{    # 動產
                            'license_plate': None,   # 車牌
                            'debtor': None,          # 債務人
                            'contract_start': None,  # 契約起始
                            'contract_end': None,    # 契約終止
                            'secured_debt': None,    # 擔保債權
                            'mortgagee': None,       # 抵押權人
                            'owner': None,           # 所有人
                            'info_data': {    # 基本資料
                                        'registration_authority': None,     # 機關
                                        'case_category': None,              # 案件類別
                                        'registration_number': None,        # 登記編號
                                        'registration_approval_date': None, # 登記核准日期
                                        'change_document_number': None,     # 變更文號
                                        'change_approval_date': None,       # 變更核准日期
                                        'cancellation_number': None,        # 註銷文號
                                        'cancellation_date': None           # 註銷日期
                                        },
                            'debtor_data': {    # 債務人資料
                                            'name': None,       # 名稱
                                            'uid':  None,       # 統編
                                            'agent_name': None, # 代理人名稱
                                            'agent_uid': None   # 代理人統編
                                            },
                            'mortgagee_data': {     # 抵押權人資料
                                            'name': None,       # 名稱
                                            'uid':  None,       # 統編
                                            'agent_name': None, # 代理人名稱
                                            'agent_uid': None   # 代理人統編
                                            },
                            'chattel_mortgage_data': {    # 動產擔保交易資料
                                                    'contract_start': None,        # 契約啟始日期
                                                    'contract_end': None,          # 契約終止日期
                                                    'target_owner_name': None,     # 標的物所有人名稱
                                                    'guarantee_amount': None,      # 擔保債權金額
                                                    'target_owner_uid': None,      # 標的物所有人統編
                                                    'property_detail': None,       # 動產明細項數
                                                    'target_location': None,       # 標的物所在地
                                                    'is_maximum': None,            # 是否最高限顏
                                                    'is_floating': None,           # 是否為浮動擔保
                                                    'target_type': None            # 標的物種類
                                                    }
                            }],
                    }]
        return data_list

    def check_phone(self, phone):
        if phone[:2] != '09' or len(phone) != 10:
            return True
        try:
            int(phone)
        except:
            return True
        return False

    def process(self, request):
        result = {'status': 'NG'}
        try:
            params = request.POST
            ptoken = params.get('ptoken', None)
            is_call = params.get('is_call', 0)
            is_call = int(is_call)
            call_end = params.get('call', 0)
            call_end = int(call_end)
            thunderbolt = params.get('thunderbolt', True)
            #! 控制蛋
            role = check_role(request)
            egg = True if thunderbolt and role in [0, 1] else False
            if not ptoken:
                result['msg'] = '請輸入參數'
                return result
            pdata = aes_decrypt(ptoken)
            if self.check_phone(pdata):
                result['msg'] = '參數錯誤'
                return result
            phone = pdata if egg else pdata[:4] + '******'
            phone_egg = pdata[:4] + '******'
            if not pdata:
                data = {
                        'phone': '',
                        'tag': [],
                        'persons': self.none_data(phone),
                        'transfer_out_type': 0,
                        'handle_type': 0
                        }
                result['status'] = 'OK'
                result['msg'] = '成功傳送資料'
                result['data'] = data
                return result

            uid_tag_type = {0: '未知', 1: '本國人', 2: '外國人或無國籍人', 3: '取得國籍之外國人', 4: '原無戶籍國民', 5: '原港澳人民', 6: '原大陸人民'}
            gender_type = {0: '未知', 1: '男', 2: '女'}
            time_now = timezone.now()
            try:
                #! 測試用
                if is_test:
                    role = 1
                tag = []
                transfer_out = 0
                handle_type = 0
                check_transfer_out = False
                # if role in [1, 2]:
                if role == 0:
                    record_data = TIRecord.objects.filter(phone=phone_egg if egg else phone, ptoken=ptoken).order_by('-update_time')
                else:
                    company_id = CompanyUserMapping.objects.filter(user=self.user, is_valid=1)[0].company_id
                    users = CompanyUserMapping.objects.filter(company_id=company_id, is_valid=1)
                    user_list = []
                    for i in users:
                        if not i.user_id in users:
                            user_list.append(i.user_id)
                    # user_list.append(self.user.id)
                    # print(phone, ptoken, user_list)
                    record_data = TIRecord.objects.filter(user_id__in=user_list, phone=phone_egg if egg else phone, ptoken=ptoken).order_by('-update_time')

                tag = record_data[0].tag if record_data else []
                record_list = []
                for i in record_data:
                    if not i.id in record_list:
                        record_list.append(i.id)
                record_datas = History.objects.filter(ti_record_id__in=record_list).order_by('-create_time')
                transfer_out = record_datas[0].transfer_out_type if record_datas else 0
                handle_type = record_datas[0].handle_type if record_datas else 0
                # check_transfer_out = True if record_datas else False
                tag = record_datas[0].ti_record.tag if record_datas else []
                if role != 0:
                    records, _ = TIRecord.objects.get_or_create(user=self.user, phone=phone_egg if egg else phone, ptoken=ptoken)
                    #! tag 覆蓋
                    records.tag = tag
                    records.is_read = 1
                # transfer_out_datas = History.objects.filter(ti_record_id=records.id).order_by('-create_time')
                # transfer_out_data = transfer_out_datas[0].transfer_out_type if transfer_out_datas else 0
                # handle_type_data = transfer_out_datas[0].handle_type if transfer_out_datas else 0
                data = {
                        'phone': phone,
                        'tag': tag,
                        'transfer_out_type': transfer_out,
                        'handle_type': handle_type,
                        'persons': self.none_data(phone),
                        'history': []
                        }
            except Exception as e:
                print(e, 'exception in line', sys.exc_info()[2].tb_lineno)
                if 'returned more than one TIRecord' in str(e):
                    data = {
                            'phone': '',
                            'tag': [],
                            'persons': self.none_data(phone),
                            'transfer_out_type': 0,
                            'handle_type': 0
                            }
                    result['status'] = 'OK'
                    result['msg'] = '成功傳送資料'
                    result['data'] = data
                    return result
                data = {
                    'phone': phone,
                    'tag': [],
                    'persons': self.none_data(phone),
                    'history': [],
                    'transfer_out_type': 0,
                    'handle_type': 0
                    }
                # with transaction.atomic():
                #     records = TIRecord.objects.create(user=self.user, phone=phone, ptoken=ptoken)

            #! 土建資料處理
            lbkey_regno_list = []
            lkey_regno_sql = ''
            bkey_regno_sql = ''
            u_lkey_regno_dict = {}
            u_bkey_regno_dict = {}
            u_lbkey_dict = {}
            #! 沒比對到謄本處理
            try:
                sql = f'''SELECT a.id, a.b4, c.c3, c.c4, c.c5 FROM telem.i_search_babaco a
                        left join telem.i_search_citron c on a.b4=c.c2 and c.c6=1
                        where a.b5="{ptoken}" and a.b7=1'''
                land_datas = Babaco.objects.raw(sql)
                for i in land_datas:
                    if not i.c4 or not i.c5:
                        continue
                    lbkey_regno = i.c4 + ';' + i.c5
                    udata = i.b4
                    lbtype = i.c3
                    if not lbkey_regno in lbkey_regno_list:
                        lbkey_regno_list.append(lbkey_regno)
                        if lbtype == 'L':
                            lkey_regno_sql += f'or (lbkey="{i.c4}" and regno="{i.c5}")' if lkey_regno_sql else f'(lbkey="{i.c4}" and regno="{i.c5}")'
                        elif lbtype =='B':
                            bkey_regno_sql += f'or (lbkey="{i.c4}" and regno="{i.c5}")' if bkey_regno_sql else f'(lbkey="{i.c4}" and regno="{i.c5}")'
                        u_lbkey_dict[lbkey_regno] = udata
            except Exception as e:
                print(e)
                data = {
                        'phone': '',
                        'tag': [],
                        'persons': self.none_data(phone),
                        'transfer_out_type': 0,
                        'handle_type': 0
                        }
                result['status'] = 'OK'
                result['msg'] = '成功傳送資料'
                result['data'] = data
                return result

            #! 土地 
            if lkey_regno_sql:
                sql = f'''SELECT a.id, a.lbkey, a.regno, b.city_name, b.area_name, b.region_name, b.land_notice_value, a.reg_date_str, a.reg_reason_str, a.right_str,
                    a.shared_size, a.creditors_rights
                    FROM diablo.t_search_lkeyregnolist a
                    left join diablo.t_search_lmlandlist b on a.lbkey=b.lkey
                    where {lkey_regno_sql} and a.remove_time is null and a.is_valid=1;'''
                land_datas = LkeyRegnoList.objects.raw(sql)
                for i in land_datas:
                    lbkey = i.lbkey
                    regno = i.regno
                    lbkey_regno = lbkey + ';' + regno
                    udata = u_lbkey_dict[lbkey_regno]
                    shared_size = round(float(i.shared_size) * 0.3025, 2) if i.shared_size else 0
                    lkey_data = {
                                'city_name': i.city_name,
                                'area_name': i.area_name,
                                'region_name': i.region_name,
                                'lno': lbkey[10:],
                                'regno': regno,
                                'reg_date_str': i.reg_date_str,
                                'reg_reason_str': i.reg_reason_str,
                                'right_str': i.right_str,
                                'shared_size': shared_size,
                                'present_value': i.land_notice_value,
                                'owner_data': json.loads(i.creditors_rights.replace("'", '"')) if i.creditors_rights else None
                                }
                    if udata in u_lkey_regno_dict:
                        u_lkey_regno_dict[udata].append(lkey_data)
                    else:
                        u_lkey_regno_dict[udata] = [lkey_data]

            #! 建物
            if bkey_regno_sql:
                sql = f'''SELECT a.id, a.lbkey, a.regno, b.city_name, b.area_name, b.region_name, b.community_name, b.door, b.build_size, b.main_size, b.attach_size,
                    b.public_size, b.parking_size, b.total_level,
                    a.reg_date_str, a.reg_reason_str,
                    a.restricted_reason, a.creditors_rights
                    FROM diablo.t_search_bkeyregnolist a
                    left join diablo.t_search_bmbuildlist b on a.lbkey=b.bkey
                    where {bkey_regno_sql} and a.remove_time is null and a.is_valid=1;'''
                bkey_datas = BkeyRegnoList.objects.raw(sql)
                for i in bkey_datas:
                    lbkey = i.lbkey
                    regno = i.regno
                    lbkey_regno = lbkey + ';' + regno
                    udata = u_lbkey_dict[lbkey_regno]
                    totalArea = round(float(i.build_size) * 0.3025, 2)
                    masterArea = round(float(i.main_size) * 0.3025, 2)
                    slaveArea = round(float(i.attach_size) * 0.3025, 2)
                    publicArea = round(float(i.public_size) * 0.3025, 2)
                    carArea = round(float(i.parking_size) * 0.3025, 2)
                    bkey_data = {
                                'city': i.city_name,
                                'district': i.area_name,
                                'segment': i.region_name,
                                'buildName': i.community_name,
                                'houseNumber': i.door,
                                'otherAuthority': json.loads(i.creditors_rights.replace("'", '"')) if i.creditors_rights else None,
                                'regno': regno,
                                'reg_date_str': i.reg_date_str,
                                'reg_reason_str': i.reg_reason_str,
                                'totalArea': totalArea,
                                'masterArea': masterArea,
                                'slaveArea': slaveArea,
                                'publicArea': publicArea,
                                'carArea': carArea,
                                'floorNum': i.total_level,
                                'ownershipMark': i.restricted_reason,
                                'buildNum': lbkey[10:]
                                }
                    if udata in u_bkey_regno_dict:
                        u_bkey_regno_dict[udata].append(bkey_data)
                    else:
                        u_bkey_regno_dict[udata] = [bkey_data]

            #! 整合
            sql = f'''SELECT * FROM telem.i_search_babaco a
                    left join telem.i_search_abiu b on a.b4=b.a8
                    left join telem.i_search_damson d on a.b4=d.d2 and d.d5=1
                    where a.b5="{ptoken}" and a.b7=1'''
            total_datas = Babaco.objects.raw(sql)
            if total_datas:
                data['persons'] = []
            for i in total_datas:
                udata = i.b4
                land_data = u_lkey_regno_dict[udata] if udata in u_lkey_regno_dict else []
                build_data = u_bkey_regno_dict[udata] if udata in u_bkey_regno_dict else []
                bday = aes_decrypt(i.a5)
                bdata = aes_decrypt(i.a10) if i.a10 else ''
                property_datas = json.loads(aes_decrypt(i.d4)) if i.d4 and egg else json.loads(i.d3) if i.d3 else []
                property_data = []
                #! 從動保拉基本資料
                uid = None
                name = None
                address = None
                for datas in property_datas:
                    uid = datas['債務人資料']['統編'] if datas['債務人資料']['統編'] else None
                    name = datas['債務人資料']['名稱'] if datas['債務人資料']['名稱'] else None
                    address = datas['動產擔保交易資料']['標的物所在地'] if datas['動產擔保交易資料']['標的物所在地'] else None
                    property_data.append({
                                        'license_plate': datas['車牌'] if '車牌' in datas else None,   # 車牌
                                        'debtor': datas['債務人名稱'],          # 債務人
                                        'contract_start': datas['動產擔保交易資料']['契約啟始日期'],  # 契約起始
                                        'contract_end': datas['動產擔保交易資料']['契約終止日期'],    # 契約終止
                                        'secured_debt': datas['動產擔保交易資料']['擔保債權金額'],    # 擔保債權
                                        'mortgagee': datas['抵押權人名稱'],       # 抵押權人
                                        'owner': datas['債務人名稱'],           # 所有人
                                        'info_data': {    # 基本資料
                                                    'registration_authority': datas['案件基本資料']['登記機關'],     # 機關
                                                    'case_category': datas['案件基本資料']['案件類別'],              # 案件類別
                                                    'registration_number': datas['案件基本資料']['登記編號'],        # 登記編號
                                                    'registration_approval_date': datas['案件基本資料']['登記核准日期'], # 登記核准日期
                                                    'change_document_number': datas['案件基本資料']['變更文號'],     # 變更文號
                                                    'change_approval_date': datas['案件基本資料']['變更核准日期'],       # 變更核准日期
                                                    'cancellation_number': datas['案件基本資料']['註銷文號'],        # 註銷文號
                                                    'cancellation_date': datas['案件基本資料']['註銷日期']           # 註銷日期
                                                    },
                                        'debtor_data': {    # 債務人資料
                                                        'name': datas['債務人資料']['名稱'],       # 名稱
                                                        'uid':  datas['債務人資料']['統編'],       # 統編
                                                        'agent_name': datas['債務人資料']['代理人名稱'], # 代理人名稱
                                                        'agent_uid': datas['債務人資料']['代理人統編']   # 代理人統編
                                                        },
                                        'mortgagee_data': {     # 抵押權人資料
                                                        'name': datas['抵押權人資料']['名稱'],       # 名稱
                                                        'uid':  datas['抵押權人資料']['統編'],       # 統編
                                                        'agent_name': datas['抵押權人資料']['代理人名稱'], # 代理人名稱
                                                        'agent_uid': datas['抵押權人資料']['代理人統編']   # 代理人統編
                                                        },
                                        'chattel_mortgage_data': {    # 動產擔保交易資料
                                                                'contract_start': datas['動產擔保交易資料']['契約啟始日期'],        # 契約啟始日期
                                                                'contract_end': datas['動產擔保交易資料']['契約終止日期'],          # 契約終止日期
                                                                'target_owner_name': datas['動產擔保交易資料']['標的物所有人名稱'],     # 標的物所有人名稱
                                                                'guarantee_amount': datas['動產擔保交易資料']['擔保債權金額'],      # 擔保債權金額
                                                                'target_owner_uid': datas['動產擔保交易資料']['標的物所有人統編'],      # 標的物所有人統編
                                                                'property_detail': datas['動產擔保交易資料']['動產明細項數'],       # 動產明細項數
                                                                'target_location': datas['動產擔保交易資料']['標的物所在地'],       # 標的物所在地
                                                                'is_maximum': datas['動產擔保交易資料']['是否最高限額'],            # 是否最高限顏
                                                                'is_floating': datas['動產擔保交易資料']['是否為浮動擔保'],           # 是否為浮動擔保
                                                                'target_type': datas['動產擔保交易資料']['標的物種類']            # 標的物種類
                                                                }
                                        })
                data_dict = {
                            'user_info': {
                                        'uid': aes_decrypt(i.a8) if i.a8 and egg else aes_decrypt(i.b4) if i.b4 and egg else i.a1 if i.a1 else i.b1 if i.b1 else uid,
                                        'name': aes_decrypt(i.a9) if i.a9 and egg else i.a2 if i.a2 else name,
                                        'bday': ''.join([bdata[:3], '年', bdata[3:]]) if bdata and egg else ''.join([bday[:3], '年', bday[3:]]) if bday else None,
                                        'gender': gender_type[i.a7] if i.a7 else None,
                                        'cell_phone': phone,
                                        'phone': i.b3,
                                        'addr': i.a3 if i.a3 else address,
                                        'address_current': i.a4,
                                        'uid_tag': uid_tag_type[i.a6] if i.a6 else None
                                        },
                            'land': land_data,
                            'build': build_data,
                            'Property': property_data
                            }
                data['persons'].append(data_dict)
            if data['persons'] == []:
                data['persons'] = self.none_data(phone)

            #! 處理歷史紀錄
            history_data = []
            history_datas = 0
            if role != 0:
                history_datas = History.objects.filter(ti_record=records).order_by('-create_time')
            check_update_time = False
            # 沒歷史紀錄和別人轉出
            if not history_datas and role != 0:
                check_update_time = True
                History.objects.create(ti_record=records, handle_type=handle_type, memo=f'{self.user.first_name} 首次匯入 \n {datetime.strftime(datetime.now(), "%Y %m %d %H:%M:%S")}',
                                        transfer_out_type=transfer_out)
            #! 接電話紀錄
            if is_call:
                History.objects.create(ti_record=records, handle_type=handle_type, memo=f'{self.user.first_name} 接聽電話 \n {datetime.strftime(datetime.now(), "%Y %m %d %H:%M:%S")}',
                                        transfer_out_type=transfer_out)
            #! 結束電話紀錄
            if call_end:
                History.objects.create(ti_record=records, handle_type=handle_type, memo=f'{self.user.first_name} 結束電話 \n {datetime.strftime(datetime.now(), "%Y %m %d %H:%M:%S")}',
                                        transfer_out_type=transfer_out)
            # history_datas = History.objects.filter(ptoken=ptoken).order_by('-create_time')
            for i in record_datas:
                if not i.memo:
                    continue
                history_data.append(i.memo)
            #! 更新時間
            if role != 0:
                if check_update_time or is_call or call_end:
                    records.update_time = time_now
                with transaction.atomic():
                    records.save()
            data['history'] = history_data
            result['status'] = 'OK'
            result['msg'] = '成功傳送資料'
            result['data'] = data
        except Exception as e:
            print(e, 'exception in line', sys.exc_info()[2].tb_lineno)
            result['msg'] = '發生不可預期的錯誤，請聯繫官方客服'
        return result

    @extend_schema(
        summary='取電話紀錄',
        description='取電話紀錄',
        request=SaveRecordSerializer,
        responses={
            200: OpenApiResponse(description='ok'),
            401: OpenApiResponse(description='身分認證失敗'),
        },
    )

    def post(self, request):
        self.user = User.objects.get(username=request.user.get_username())
        logger.info('取電話紀錄')
        time_start = time.perf_counter()
        result = self.process(request)
        time_end = time.perf_counter()
        logger.info(f'花費時間：{time_end - time_start}秒')
        if result['status'] == 'NG':
            return HttpResponseBadRequest(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")
        return HttpResponse(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")

class GetTirecordView(APIView):
    TAG = "[GetTirecord]"

    authentication_classes = (CsrfExemptSessionAuthentication, )

    def my_own_data(self, ptoken_list):
        datas = TIRecord.objects.filter(user_id=self.user.id, ptoken__in=ptoken_list)
        # print(ptoken_list)
        my_data_dict = {}
        for data in datas:
            my_data_dict[data.ptoken] = {
                                    'create_time': timezone.localtime(data.create_time).strftime("%Y-%m-%d") if data.create_time else '', 
                                    'update_time': timezone.localtime(data.update_time).strftime("%Y-%m-%d") if data.update_time else '',
                                    'read': 1 if data.is_read else 0
                                    }
        return my_data_dict

    def handle_sys_transfer(self, auto_transfer_dict):
        user_data = CompanyUserMapping.objects.filter(user_id=self.user.id, is_valid=1)[0]
        user_datas = CompanyUserMapping.objects.filter(company_id=user_data.company_id, is_valid=1)
        user_list = []
        for i in user_datas:
            if (i.is_admin or i.is_manager) and not i.user in user_list:
                user_list.append(i.user)

        for k, v in auto_transfer_dict.items():
            phone = aes_decrypt(k)[:4] + '******'
            try:
                for s, i in enumerate(user_list):
                    records = TIRecord.objects.create(user=i, phone=phone, ptoken=k, tag=v['tag'], update_time=self.time_now)
                    if s == 0:
                        History.objects.create(ti_record=records, handle_type=v['handle_type'], memo=f'系統自動從operator轉出 \n {self.time_now_str}', transfer_out_type=2)
                    History.objects.create(ti_record=records, handle_type=v['handle_type'], memo=f'{i.first_name} 首次匯入 \n {self.time_now_str}', transfer_out_type=2)
            except Exception as e:
                # print(e)
                pass

    def process(self, request):
        result = {'status': 'NG','msg':'發生不可預期的錯誤，請聯繫官方客服'}
        role = check_role(request)
        date_7 = datetime.strftime(datetime.now() - timedelta(days=7), "%Y-%m-%d")
        date_10 = datetime.strftime(datetime.now() - timedelta(days=10), "%Y-%m-%d")
        self.time_now = timezone.now()
        self.time_now_str = datetime.strftime((datetime.now()), "%Y %m %d %H:%M:%S")
        #! 控制蛋
        egg = True if role in [0, 1] else False
        # role = 1
        #! 測試用
        # is_test = True
        if is_test:
            role = 2
        if role in [0, 1, 2]:
            try:
                if role == 0:
                    sql = f'''
                            SELECT 1 as `id`, f.a2, f.a9, c.phone, c.tag, c.ptoken, d.handle_type, d.transfer_out_type
                            FROM telem.i_search_tirecord c
                            left join telem.i_search_history d on c.id = d.ti_record_id
                            left join telem.i_search_babaco e on e.b5=c.ptoken
                            left join telem.i_search_abiu f on e.b4=f.a8
                            where d.transfer_out_type not in (3,4)
                            and c.id is not null and d.ti_record_id is not null 
                            order by c.update_time desc
                            '''
                else:
                    sql = f'''
                            SELECT 1 as `id`, f.a2, f.a9, c.phone, c.tag, c.ptoken, d.handle_type, d.transfer_out_type, a.user_id
                            FROM telem.users_companyusermapping a
                            left join telem.users_companyusermapping b on a.company_id = b.company_id
                            left join (SELECT * FROM telem.i_search_tirecord where update_time in (select update_time from telem.i_search_tirecord group by ptoken)) c on c.user_id = b.user_id
                            left join (SELECT * FROM telem.i_search_history where create_time in (select max(create_time) from telem.i_search_history group by ti_record_id)) d on c.id = d.ti_record_id
                            left join telem.i_search_babaco e on e.b5=c.ptoken
                            left join telem.i_search_abiu f on e.b4=f.a8
                            where ((c.user_id={self.user.id} and d.transfer_out_type not in (3,4)) or (a.user_id={self.user.id} and (d.transfer_out_type=1 or (d.transfer_out_type=2 and DATE_SUB(NOW(), INTERVAL 10 DAY) >= c.update_time))))
                            and c.id is not null and d.ti_record_id is not null 
                            order by c.update_time desc
                            '''
                # print(sql)
                data_list = []
                total_datas = TIRecord.objects.raw(sql)
                name_mapping = {}
                ptoken_list = []
                auto_transfer_dict = {}
                for i in total_datas:
                    ptoken = i.ptoken
                    if i.transfer_out_type == 2 and not ptoken in auto_transfer_dict:
                        auto_transfer_dict[ptoken] = {
                                            'tag': i.tag,
                                            'handle_type': i.handle_type
                                            }
                    if not ptoken in ptoken_list:
                        ptoken_list.append(ptoken)
                    if not i.a2:
                        continue
                    name = aes_decrypt(i.a9) if egg and i.a9 else i.a2
                    ptoken = i.ptoken
                    if ptoken in name_mapping and not name in name_mapping[ptoken]:
                        name_mapping[ptoken].append(name)
                    else:
                        name_mapping[ptoken] = [name]
                check_ptoken = []
                #! 處理自己是否已讀和時間
                my_data_dict = self.my_own_data(ptoken_list) if ptoken_list else {}
                #! 處理 operator 超過10天的待追蹤
                if auto_transfer_dict and role != 0:
                    self.handle_sys_transfer(auto_transfer_dict)
                for i in total_datas:
                    ptoken = i.ptoken
                    if not ptoken in check_ptoken:
                        check_ptoken.append(ptoken)
                        data_list.append({
                            'name_list': name_mapping[ptoken] if ptoken in name_mapping else [],
                            'phone': aes_decrypt(ptoken) if egg and ptoken else i.phone,
                            'transfer_out_type': i.transfer_out_type,
                            'handle_type': i.handle_type,
                            'tag': i.tag,
                            'ptoken': ptoken,
                            'read': 1 if role == 0 else my_data_dict[ptoken]['read'] if ptoken in my_data_dict else 0,
                            'remind': 0,
                            # 'remark': ''
                        })

                    result['status'] = 'OK'
                    result['msg'] = '成功傳送資料'
                    result['data'] = data_list if data_list else []

                return result

            except Exception as e:
                logger.info(f"{e} exception in line {sys.exc_info()[2].tb_lineno}")
                # print(e, 'exception in line', sys.exc_info()[2].tb_lineno)
        elif role == 3:
            try:
                sql = f'''
                        SELECT 1 as `id`, a.a2, c.phone, c.tag, c.ptoken, c.create_time, c.update_time, d.handle_type, d.transfer_out_type
                        FROM telem.i_search_tirecord t
                        left join (SELECT * FROM telem.i_search_tirecord where update_time in (select max(update_time) from telem.i_search_tirecord group by ptoken)) c on c.ptoken = t.ptoken
                        left join telem.i_search_babaco b on b.b5=c.ptoken
                        left join telem.i_search_abiu a on b.b4=a.a8
                        left join (SELECT * FROM telem.i_search_history where create_time in (select max(create_time) from telem.i_search_history group by ti_record_id)) d on c.id = d.ti_record_id
                        where t.user_id={self.user.id} and (c.update_time > curdate() or d.transfer_out_type=2) and d.transfer_out_type not in (1,3,4)
                        order by c.update_time desc
                        '''
                # print(sql)
                data_list = []
                total_datas = TIRecord.objects.raw(sql)
                name_mapping = {}
                ptoken_list = []
                for i in total_datas:
                    ptoken = i.ptoken
                    if not ptoken in ptoken_list:
                        ptoken_list.append(ptoken)
                    if not i.a2:
                        continue
                    name = aes_decrypt(i.a9) if egg and i.a9 else i.a2
                    if ptoken in name_mapping and not name in name_mapping[ptoken]:
                        name_mapping[ptoken].append(name)
                    else:
                        name_mapping[ptoken] = [name]
                #! 處理自己是否已讀和時間
                my_data_dict = self.my_own_data(ptoken_list)
                check_ptoken = []
                for i in total_datas:
                    ptoken = i.ptoken
                    if not ptoken in check_ptoken:
                        check_ptoken.append(ptoken)
                        data_date = my_data_dict[ptoken]['update_time'] if my_data_dict[ptoken]['update_time'] else my_data_dict[ptoken]['create_time']
                        remind = 2 if data_date <= date_10 else 1 if data_date <= date_7 else 0
                        remark = '更新於10天以前' if remind == 2 else '更新於7天以前，資料於第10天轉出給主管' if remind == 1 else ''
                        #! 超過10天不顯示
                        if remind == 2:
                            continue
                        data_list.append({
                            'name_list': name_mapping[ptoken] if ptoken in name_mapping else [],
                            'phone': aes_decrypt(ptoken) if egg and ptoken else i.phone,
                            'transfer_out_type': i.transfer_out_type,
                            'handle_type': i.handle_type,
                            'tag': i.tag,
                            'ptoken': ptoken,
                            'read': my_data_dict[ptoken]['read'] if ptoken in my_data_dict else 0,
                            'remind': remind,
                            # 'remark': remark
                        })

                if data_list:
                    result['status'] = 'OK'
                    result['msg'] = '成功傳送資料'
                    result['data'] = data_list
                    return result
                else:
                    result['status'] = 'OK'
                    result['msg'] = '成功傳送資料'
                    result['data'] = []
                    return result
            except Exception as e:
                logger.info(f"{e} exception in line {sys.exc_info()[2].tb_lineno}")
                # print(e, 'exception in line', sys.exc_info()[2].tb_lineno)
        elif role == 4:
            result['msg'] = '無權限'

        return result

    @extend_schema(
        summary='取歷史紀錄',
        description='取歷史紀錄',
        request=TirecordSerializer,
        responses={
            200: OpenApiResponse(description='ok'),
            401: OpenApiResponse(description='身分認證失敗'),
        },
    )

    def get(self, request):
        self.user = User.objects.get(username=request.user.get_username())
        logger.info('取歷史紀錄')
        time_start = time.perf_counter()
        result = self.process(request)
        time_end = time.perf_counter()
        logger.info(f'花費時間：{time_end - time_start}秒')
        if result['status'] == 'NG':
            return HttpResponseBadRequest(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")
        return HttpResponse(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")

class AddRemark(APIView):
    TAG = "[AddRemark]"

    authentication_classes = (CsrfExemptSessionAuthentication, )

    def process(self, request):
        result = {'status': 'NG'}
        try:
            params = request.POST
            ptoken = params.get('ptoken', None)
            tag_list = params.get('tag_list', None)
            handle_type = params.get('handle_type', 0)
            transfer_out = params.get('transfer_out_type', 0)
            memo_text = params.get('memo', None)
            utoken = params.get('utoken', None)
            if not ptoken:
                result['msg'] = '請輸入參數'
                return result
            role = check_role(request)
            handle_type_dict = {0: '未設定', 1: '聯繫', 2: '更新進度', 3: '結案'}
            transfer_out_type = {0: '未設定', 1: '有意願購買,轉主管', 2: '加入待追蹤名單', 3: '放棄,電訪三次找不到人', 4: '放棄,無意願購買'}
            time_now = datetime.now()
            try:
                ti = TIRecord.objects.get(user=self.user, ptoken=ptoken)
            except:
                result['msg'] = '最高身份，請勿改紀錄'
                return result
            if tag_list and tag_list != "[]":
                tag_list = json.loads(tag_list)
                ti.tag = tag_list
            else:
                tag_list = []

            tag_text = '、'.join(tag_list) if tag_list else ''
            handle_type = int(handle_type)
            handle_type_str = ' ' + handle_type_dict[int(handle_type)] if handle_type else ''
            memo_text = ' 接洽紀錄：' + memo_text if memo_text else ''
            newest_memo, ti_update_unread = company_newest_history(ptoken, self.user.id)
            try:
                transfer_out_str = transfer_out_type[int(transfer_out)]
            except:
                transfer_out = 0
                transfer_out_str = transfer_out_type[0]
            transfer_out = int(transfer_out)

            if role == 0:
                result['msg'] = '不好吧'
                return result
            elif (role in [1, 2] and transfer_out == 1) or (role == 3 and handle_type == 3):
                result['msg'] = '操作錯誤'
                return result

            memo = ''
            add_check = ''
            if tag_text and not f' 標籤 {tag_text}' in newest_memo:
                memo += f'{self.user.first_name} 標籤 {tag_text}'
                add_check = '\n'
            if transfer_out and not f' 轉出 {transfer_out_str}' in newest_memo:
                memo += f'{add_check}{self.user.first_name} 轉出 {transfer_out_str}'
                add_check = '\n'
            if memo_text or handle_type and not f'{handle_type_str}{memo_text}' in newest_memo:
                memo += f'{add_check}{self.user.first_name}{handle_type_str}{memo_text}'
                add_check = '\n'
            if tag_text or transfer_out or memo_text or handle_type:
                memo += f'{add_check}{datetime.strftime(time_now, "%Y %m %d %H:%M:%S")}'
                ti_update_unread.update(is_read=0)

            with transaction.atomic():
                ti.update_time = time_now
                ti.save()
                History.objects.create(ti_record=ti, handle_type=handle_type, memo=memo, transfer_out_type=transfer_out)

            result['status'] = 'OK'
            result['msg'] = '成功儲存資料'
        except Exception as e:
            print(e, 'exception in line', sys.exc_info()[2].tb_lineno)
            result['msg'] = '發生不可預期的錯誤，請聯繫官方客服'
        return result

    @extend_schema(
        summary='新增紀錄',
        description='新增紀錄',
        request=AddRemarkSerializer,
        responses={
            200: OpenApiResponse(description='ok'),
            401: OpenApiResponse(description='身分認證失敗'),
        },
    )

    def post(self, request):
        self.user = User.objects.get(username=request.user.get_username())
        logger.info('新增紀錄')
        time_start = time.perf_counter()
        result = self.process(request)
        time_end = time.perf_counter()
        logger.info(f'花費時間：{time_end - time_start}秒')
        if result['status'] == 'NG':
            return HttpResponseBadRequest(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")
        return HttpResponse(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")

# processing csv/excel file lbkeys
class UploadFileView(LoginRequiredMixin, TemplateView):
    TAG = '[UploadFileView]'
    template_name = "upload_file.html"

    def get_context_data(self, **kwargs):
        context = super(UploadFileView, self).get_context_data(**kwargs)
        return context

    def read_csv_lbkeys(self, upload_file):
        filename = upload_file.name
        response = None
        datas = list()
        try:
            csv_file = StringIO(upload_file.read().decode())
            reader = csv.reader(csv_file)
            headers = next(reader)
            for row in reader:
                if row and len(row)==2:
                    datas.append(row)
            if len(datas) == 0:
                response = {'status': 'NG', 'msg': 'No valid data or incorrect format in csv file uploaded'}
        except Exception as error:
            logger.error('{}, {}, failed: {}, {}'.format(self.TAG, filename, error, traceback.print_exc()))
            response = {'status': 'NG', 'msg': '{}'.format(error), 'file': filename}
        return datas, response

    def read_excel_lbkeys(self, upload_file):
        filename = upload_file.name
        response = None
        datas = list()
        try:
            wb = load_workbook(filename=BytesIO(upload_file.read()))
            #  read from Excel file to get old infos
            ws = wb.worksheets[0]
            if ws.max_row >= 1:
                for i in range(2, ws.max_row + 1):
                    if not ws.cell(row=i, column=1).value:
                        break
                    data = []
                    for j in range(1, ws.max_column + 1):
                        data.append(str(ws.cell(row=i, column=j).value))
                    datas.append(data)
            if len(datas) == 0:
                response = {'status': 'NG', 'msg': 'No valid data or incorrect format in Excel file uploaded'}
        except Exception as error:
            logger.error('{}, {}, failed: {}, {}'.format(self.TAG, filename, error, traceback.print_exc()))
            response = {'status': 'NG', 'msg': '{}'.format(error)}
        return datas, response

    def read_xml(self, upload_file):
        filename = upload_file.name
        response = None
        datas = list()
        try:
            df = pd.read_xml(BytesIO(upload_file.read()))
            #  read from xml file to get infos
            for index, row in df.iterrows():
                data = []
                for column in row:
                    data.append(str(column))
                datas.append(data)
            if len(datas) == 0:
                response = {'status': 'NG', 'msg': 'No valid data or incorrect format in xml file uploaded'}
                # df_out = df.head(5)
                # df_out = df.iloc[:5]
                # xml_output = df_out.to_xml()
                # with open("rows_5.xml", "w") as file:
                #     file.write(xml_output)
        except Exception as error:
            logger.error('{}, {}, failed: {}, {}'.format(self.TAG, filename, error, traceback.print_exc()))
            response = {'status': 'NG', 'msg': '{}'.format(error)}
        return datas, response

    def process(self, request, upload_file, **kwargs):
        filename = upload_file.name
        try:
            #! 金鑰
            key = settings.THUNDERBOLT_KEY
            key = key.encode('utf8')
            if upload_file.name.split(".")[-1] == 'xlsx':
                datas, response = self.read_excel_lbkeys(upload_file)
            elif upload_file.name.split(".")[-1] == 'csv':
                datas, response = self.read_csv_lbkeys(upload_file)
            else:
                raise Exception("檔案類型錯誤")
            if len(datas) > 0:
                create_list = []
                for data in datas:
                    #! to do here
                    data_0 = data[0].replace(' ', '').upper()
                    data_11 = pdata_handle(data[1])
                    for data_1 in data_11:
                        if len(data_1) == 10:
                            b1 = data_0[:4] + '*****' + data_0[-1]
                            b2 = data_1[:4] + '******'
                            b4 = aes_encrypt(data_0)
                            b5 = aes_encrypt(data_1)
                            kwargs = {
                                    'b1': b1,
                                    'b2': b2,
                                    'b4': b4,
                                    'b5': b5
                                    }
                            create = Babaco(**kwargs)
                            create_list.append(create)
                with transaction.atomic():
                    Babaco.objects.bulk_create(create_list, batch_size=10000, ignore_conflicts=True)
            results = {'status': 'OK', 'msg': '檔案匯入完成'}
        except Exception as error:
            results = {'status': 'NG', 'msg': '檔案匯入發生錯誤({})，請聯繫官方客服'.format(error)}
        return results

    def profile_handle(self, c_uid):
        city_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L' ,'M', 'N', 'O', 'P', 'Q','R','S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        uid_tag_89 = ['0', '1', '2', '3', '4', '5', '6']
        uid_tag_12 = uid_tag_89[:-1]
        uid_tag_dict = {'6': 3, '7': 4, '8': 5, '9': 6}
        gender_man = ['A', 'C', '1', '8']
        gender_woman = ['B', 'D', '2', '9']
        uid_tag = 0
        gender = 0
        #* 國籍判斷
        if c_uid[0] in city_list:
            #*性別判斷
            gender = 1 if c_uid[1] in gender_man else 2 if c_uid[1] in gender_woman else 0
            if c_uid[1] in city_list:
                uid_tag = 2
            elif c_uid[1] in ['8', '9']:
                uid_tag = 2 if c_uid[2] in uid_tag_89 else uid_tag_dict[c_uid[2]] if c_uid[2] in uid_tag_dict else 0
            elif c_uid[1] in ['1', '2']:
                uid_tag = 1 if c_uid[2] in uid_tag_12 else uid_tag_dict[c_uid[2]] if c_uid[2] in uid_tag_dict else 0
        #* 第一碼為數字 -> 外國人
        elif c_uid[0] in ['0','1','2','3','4','5','6','7','8','9']:
            uid_tag = 2
        return uid_tag, gender

    def process_xml(self, request, upload_file, **kwargs):
        filename = upload_file.name
        try:
            #! 金鑰
            key = settings.THUNDERBOLT_KEY
            key = key.encode('utf8')
            if upload_file.name.split(".")[-1] != 'xml':
                raise Exception("檔案類型錯誤")
            datas, response = self.read_xml(upload_file)
            create_list = []
            data_dict = {}
            data_list = []
            have_list = []
            update_list = []
            udata_dict = {}
            udata_list = []
            udata_create = []
            now_data = datetime.strftime((datetime.now()), "%Y-%m-%d")
            if len(datas) > 0:
                count = len(datas)
                for s, data in enumerate(datas):
                    data_0 = data[0].replace(' ', '')
                    d2 = aes_encrypt(data_0)
                    if not d2 in data_list:
                        data_list.append(d2)
                    data_1 = json.loads(data[1].replace("'", "\'")) if data[1] else {}
                    if data_1:
                        data_5 = copy.deepcopy(data_1)
                        d3 = copy.deepcopy(data_1)
                        for ss, data_4 in enumerate(data_1):
                            data_6 = data_4['動產擔保交易資料']['標的物所有人名稱']
                            data_7 = data_6.split('車牌號碼:')[1].replace(')', '') if data_6 and '車牌號碼' in data_6 else None
                            data_8 = data_7[:3] + '***' if data_7 else None
                            data_9 = data_4['動產擔保交易資料']['標的物所在地']
                            if ss == 0:
                                data_2 = data_4['債務人名稱']
                                data_3 = data_2[0] + '＊＊'
                                d1 = data_0[:4] + '*****' + data_0[-1]
                            data_5[ss]['車牌'] = data_7
                            data_5[ss]['債務人資料']['統編'] = data_0
                            data_5[ss]['動產擔保交易資料']['標的物所有人統編'] = data_0

                            d3[ss]['車牌'] = data_8
                            d3[ss]['債務人資料']['統編'] = d1
                            d3[ss]['動產擔保交易資料']['標的物所有人統編'] = d1
                            d3[ss]['債務人名稱'] = data_3
                            d3[ss]['債務人資料']['名稱'] = data_3
                            d3[ss]['動產擔保交易資料']['標的物所有人名稱'] = data_3

                        d4 = aes_encrypt(json.dumps(data_5, ensure_ascii=False))
                        data_dict[d2] = {
                                'd1': d1,
                                'd2': d2,
                                'd3': d3,
                                'd4': d4
                                }
                        #! 處理基本資料
                        if not d2 in udata_list:
                            udata_list.append(d2)
                            a6, a7 = self.profile_handle(data_0)
                            a9 = aes_encrypt(data_2)
                            udata_dict[d2] = {
                                            'a1': d1,
                                            'a2': data_3,
                                            'a3': data_9,
                                            'a6': a6,
                                            'a7': a7,
                                            'a8': d2,
                                            'a9': a9,
                                            'a11': now_data
                                            }

                    if ((s+1) % 5000) == 0 or (s+1) == len(datas):
                        #! 處理更新(動保)
                        have_datas = Damson.objects.filter(d2__in=data_list)
                        time_now = timezone.now()
                        for update in have_datas:
                            have_list.append(update.d2)
                            data = data_dict[update.d2]
                            update.d3 = data['d3']
                            update.d4 = data['d4']
                            update.d7 = time_now
                            update_list.append(update)
                        last_list = list(set(data_list) - set(have_list))
                        for i in last_list:
                            kwargs = data_dict[i]
                            create = Damson(**kwargs)
                            create_list.append(create)
                        #! 基本資料
                        if udata_list:
                            profile_datas = Abiu.objects.filter(a8__in=udata_list).values('a8')
                            have_a8 = []
                            for i in profile_datas:
                                have_a8.append(i['a8'])
                            last_a8 = list(set(udata_list) - set(have_a8))
                            for i in last_a8:
                                kwargs = udata_dict[i]
                                create = Abiu(**kwargs)
                                udata_create.append(create)
                        with transaction.atomic():
                            if udata_create:
                                Abiu.objects.bulk_create(udata_create, batch_size=10000, ignore_conflicts=True)
                            if create_list:
                                Damson.objects.bulk_create(create_list, batch_size=10000, ignore_conflicts=True)
                            if update_list:
                                Damson.objects.bulk_update(update_list, batch_size=5000, fields=['d3', 'd4', 'd7'])
                        #! 清空清單
                        create_list = []
                        data_dict = {}
                        data_list = []
                        have_list = []
                        update_list = []
                        udata_dict = {}
                        udata_list = []
                        udata_create = []
                        count -= 5000
                        if count > 0:
                            logger.info(f'剩餘筆數：{count}')
                        else:
                            logger.info(f'剩餘筆數：0')
            #! 更新
            call_command(owner_handle.Command(), task_type='UP')
            call_command(handle_datas.Command(), task_type='DU', control=True)
            results = {'status': 'OK', 'msg': '檔案匯入完成'}
        except Exception as error:
            print(error, 'exception in line', sys.exc_info()[2].tb_lineno)
            results = {'status': 'NG', 'msg': '檔案匯入發生錯誤({})，請聯繫官方客服'.format(error)}
        return results

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        time_start = time.perf_counter()
        form = UploadFileForm(request.POST, request.FILES)
        if not form.is_valid():
            response_data = {'status': 'NG', 'msg': form.errors}
        else:
            upload_file = request.FILES.get('file')
            if upload_file:
                if upload_file.name.split(".")[-1] in ['xlsx', 'csv']:
                    response_data = self.process(request, upload_file, *args, **kwargs)
                    response_data['file'] = upload_file.name
                elif upload_file.name.split(".")[-1] in ['xml']:
                    response_data = self.process_xml(request, upload_file, *args, **kwargs)
                    response_data['file'] = upload_file.name
                else:
                    response_data = {'status': 'NG', 'msg': 'not excel/csv/xml file: {}'.format(upload_file.name)}
            else:
                response_data = {'status': 'NG', 'msg': 'can not get uploaded file'}
        time_end = time.perf_counter()
        logger.info(f'花費時間：{time_end - time_start}秒')
        if response_data['status'] == 'NG':
            return HttpResponseBadRequest(json.dumps(response_data, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")
        return HttpResponse(json.dumps(response_data, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")

class GetLogoView(View):
    TAG = "[GetLogoView]"

    def get(self, request):
        urls = request.build_absolute_uri('/')[:-1].strip("/")
        # urls = 'http://257.vvips.com.tw/'
        #* 預設
        company = Company.objects.get(id=1)
        logo = company.logo.url
        company_name = company.company_name
        if '.telem.com.tw' in urls or '.vvips.com.tw' in urls:
            url_t = '.telem.com.tw' if '.telem.com.tw' in urls else '.vvips.com.tw' if '.vvips.com.tw' in urls else ''
            sub_domain = urls.split(url_t)[0].split('//')[1]
            company_data = Company.objects.filter(sub_domain=sub_domain, is_valid=1)
            if company_data:
                if company_data[0].logo:
                    logo = company_data[0].logo.url
                if company_data[0].company_name:
                    company_name = company_data[0].company_name
        #* 判斷角色
        try:
            role = check_role(request)
        except:
            role = None
        result = {'logo': logo, 'company_name': company_name, 'role': role}
        return HttpResponse(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")

class DownloadPhoneNumberView(APIView):
    TAG = "[DownloadPhoneNumber]"

    authentication_classes = (CsrfExemptSessionAuthentication, )

    def 處理年紀(self, sql_conds, age_interval):
        sql_str = ''
        if age_interval and age_interval != '[]':
            age_interval = json.loads(age_interval)
            templist = list()
            age_interval_dict = {'20-29': '((left(now(),4)-a14) between 20 and 29)', '30-39': '((left(now(),4)-a14) between 30 and 39)',
                                '40-49': '((left(now(),4)-a14) between 40 and 49)', '50-59': '((left(now(),4)-a14) between 50 and 59)',
                                '60-69': '((left(now(),4)-a14) between 60 and 69)', '70-79': '((left(now(),4)-a14) between 70 and 79)'}
            for item in age_interval:
                templist.append(age_interval_dict[item]) 

            sql_str = "("+ ' or '.join(templist) +")"
            sql_conds.append(sql_str)

    def 處理性別(self, sql_conds, gender):
        sql_str = ''
        if gender and gender != '[]':
            gender = json.loads(gender)
            templist = list()
            gender_dict = {'男': '(a7 = 1)', '女': '(a7 = 2)'}
            for item in gender:
                templist.append(gender_dict[item])

            sql_str = "("+ ' or '.join(templist) +")"
            sql_conds.append(sql_str)

    def 處理他項標記(self, sql_conds, credits):
        sql_str = ''
        if credits and credits != '[]':
            credits = json.loads(credits)
            templist = list()
            credits_dict = {'無設定': 0, '私設': 1, '銀行二胎': 2, '銀行': 3, '租賃': 4, '公司': 5, '政府機構': 6}
            for item in credits:
                templist.append(f"(c10 = {credits_dict[item]})")

            sql_str = "("+ ' or '.join(templist) +")"
            sql_conds.append(sql_str)

    def 處理持分狀態(self, sql_conds, rights):
        sql_str = ''
        if rights and rights != '[]':
            rights = json.loads(rights)
            templist = list()
            rights_dict = {'全部':'(c9 = 1)', '持分':'(c9 = 2)', '公同共有':'(c9 = 3)'}
            for item in rights:
                templist.append(rights_dict[item]) 
            sql_str = "("+ ' or '.join(templist) +")"
            sql_conds.append(sql_str)

    def 處理土建所在地(self, sql_conds, location):
        sql_str = ''
        if location and location != '[[]]':
            location = json.loads(location)
            templist = list()
            for item in location:
                if len(item) == 2:
                    templist.append(f"(c11 ='{item[0]}' and c12 ='{item[1]}')")
                elif len(item) == 1:
                    templist.append(f"(c11 ='{item[0]}')")
            sql_str = "("+ ' or '.join(templist) +")"
            sql_conds.append(sql_str)

    def 處理他項設定時間(self, sql_conds, setting_time):
        if setting_time:
            date_now = datetime.now()
            date_3months = f'{datetime.strftime(date_now - relativedelta(months=3), "%Y-%m-%d")} <= c14'
            date_6months = f'{datetime.strftime(date_now - relativedelta(months=6), "%Y-%m-%d")} <= c14'
            date_1years = f'{datetime.strftime(date_now - relativedelta(years=1), "%Y-%m-%d")} <= c14'
            date_2years = f'{datetime.strftime(date_now - relativedelta(years=2), "%Y-%m-%d")} <= c14'
            date_years = f'{datetime.strftime(date_now - relativedelta(years=2), "%Y-%m-%d")} >= c14'
            setting_time_dict = {'近3個月': date_3months, '近6個月': date_6months, '近1年': date_1years, '近2年': date_2years, '2年以上': date_years}
            sql_conds.append(f'({setting_time_dict[setting_time]})')

    def 處理權利人類型(self, sql_conds, right_type):
        sql_str = ''
        if right_type and right_type != '[]':
            right_type = json.loads(right_type)
            templist = list()
            right_type_dict = {'政府機構': 'property_type=1', '自然人': 'property_type=2', '公司': 'property_type=3', '租賃業者': 'property_type=4',
                                '金融機構': 'property_type=5'}
            # right_type_dict = {'自然人': 'property_type=2', '公司': 'property_type=3', '租賃業者': 'property_type=4',
            #                     '金融機構': 'property_type=5'}
            for item in right_type:
                if item in right_type_dict:
                    templist.append(right_type_dict[item]) 
            sql_str = "(" + ' or '.join(templist) + ")"
            sql_conds.append(sql_str)

    def 處理權利人名稱(self, sql_conds, right_holder):
        if right_holder:
            sql_conds.append(f'mortgagee_name like "%%{right_holder}%%"')

    def 處理擔保債權金額(self, sql_conds=[], set_amount_lo=0, set_amount_up=0):
        if set_amount_lo > 0 and set_amount_up > 0:
            sql_conds.append(f'(guarantee_amount between {set_amount_lo} and {set_amount_up})')
        elif set_amount_lo > 0:
            sql_conds.append(f'guarantee_amount>={set_amount_lo}')
        elif set_amount_up > 0:
            sql_conds.append(f'guarantee_amount<={set_amount_up}')

    def 處理契約起始時間(self, sql_conds=[], set_start_time_lo='', set_start_time_up=''):
        if set_start_time_lo and set_start_time_up:
            sql_conds.append(f'(contract_start between "{set_start_time_lo}" and "{set_start_time_up}")')
        elif set_start_time_lo:
            sql_conds.append(f'contract_start>="{set_start_time_lo}"')
        elif set_start_time_up:
            sql_conds.append(f'contract_start<="{set_start_time_up}"')

    def 處理契約終止時間(self, sql_conds=[], set_end_time_lo='', set_end_time_up=''):
        if set_end_time_lo and set_end_time_up:
            sql_conds.append(f'(contract_end between "{set_end_time_lo}" and "{set_end_time_up}")')
        elif set_end_time_lo:
            sql_conds.append(f'contract_end>="{set_end_time_lo}"')
        elif set_end_time_up:
            sql_conds.append(f'contract_end<="{set_end_time_up}"')

    def process(self, request):
        result = {'status': 'NG'}
        try:
            age_interval = request.POST.get('age_interval', "[]")
            gender = request.POST.get('gender', "[]")
            credits = request.POST.get('credits', "[]")
            rights = request.POST.get('rights', "[]")
            location = request.POST.get('location', "[[]]")
            event = request.POST.get('event', '謄本')
            lbtype = request.POST.get('lbtype', None)
            u_location = request.POST.get('u_location', "[[]]")
            right_holder = request.POST.get('right_holder', '')
            right_type = request.POST.get('right_type', "[]")
            set_amount_lo = request.POST.get('set_amount_lo', 0)
            set_amount_up = request.POST.get('set_amount_up', 0)
            set_start_time_lo = request.POST.get('set_start_time_lo', '')
            set_start_time_up = request.POST.get('set_start_time_up', '')
            set_end_time_lo = request.POST.get('set_end_time_lo', '')
            set_end_time_up = request.POST.get('set_end_time_up', '')
            setting_time = request.POST.get('setting_time', '')
            del_call_phone = request.POST.get('del_call_phone', 'true')
            del_call_phone = json.loads(del_call_phone)
            del_call_phone = 1 if del_call_phone else 0
            event_type = {'謄本': 1, '動保': 2}
            event = event_type[event]
            set_amount_lo = int(set_amount_lo)
            set_amount_up = int(set_amount_up)
            if right_type == '[""]':
                right_type = "[]"
            # set_start_time_lo = json.load(set_start_time_lo) if set_start_time_lo else ''
            # set_start_time_up = json.load(set_start_time_up) if set_start_time_up else ''
            # set_end_time_lo = json.load(set_end_time_lo) if set_end_time_lo else ''
            # set_end_time_up = json.load(set_end_time_up) if set_end_time_up else ''

            if location == '[[]]' and event != 2:
                result['msg'] = '縣市不能為空'
                return result
            datetime = timezone.now()
            try:
                company = CompanyUserMapping.objects.filter(user_id=self.user.id, is_valid=1)[0].company
            except:
                company = None
            sql_conds = list()

            self.處理年紀(sql_conds, age_interval)
            self.處理性別(sql_conds, gender)
            if event == 1:
                self.處理他項標記(sql_conds, credits)
                self.處理持分狀態(sql_conds, rights)
                self.處理土建所在地(sql_conds, location)
                self.處理他項設定時間(sql_conds, setting_time)
            elif event == 2:
                self.處理權利人類型(sql_conds, right_type)
                self.處理權利人名稱(sql_conds, right_holder)
                self.處理擔保債權金額(sql_conds, set_amount_lo, set_amount_up)
                self.處理契約起始時間(sql_conds, set_start_time_lo, set_start_time_up)
                self.處理契約終止時間(sql_conds, set_end_time_lo, set_end_time_up)
            if sql_conds:
                sql_list = ['']
                sql_list.extend(sql_conds)
                sql_mid = ' AND '.join(sql_list)
            else:
                sql_mid = []
            if del_call_phone:
                column_str = ', t.user_id'
                select_str = 'left join telem.i_search_tirecord t on b.b5=t.ptoken'
                company_id = CompanyUserMapping.objects.filter(user_id=self.user.id, is_valid=1)[0].company_id
                user_datas = CompanyUserMapping.objects.filter(company_id=company_id, is_valid=1).values('user_id')
                user_list = []
                for i in user_datas:
                    user_id = i['user_id']
                    if not user_id in user_list:
                        user_list.append(user_id)
            else:
                column_str = ''
                select_str = ''
            # print(sql_mid)
            if event == 1:
                sql = f'''
                    SELECT c.id, b5 as ptoken{column_str} FROM telem.i_search_citron c
                    left join telem.i_search_abiu a on c.c2=a.a8
                    left join telem.i_search_babaco b on a.a8=b.b4
                    {select_str}
                    where b4 is not null and b7=1{sql_mid}
                    group by b5
                    '''
            elif event == 2 and (right_holder or right_type != '[]' or set_amount_lo or set_amount_up or set_start_time_lo or set_start_time_up 
                                or set_end_time_lo or set_end_time_up):
                sql = f'''
                    SELECT d.id, b5 as ptoken{column_str} FROM telem.i_search_damsonquery dq
					left join telem.i_search_damson d on d.id=dq.damson_id
                    left join telem.i_search_abiu a on d.d2=a.a8
                    left join telem.i_search_babaco b on a.a8=b.b4
                    {select_str}
                    where b4 is not null and b7=1{sql_mid}
                    group by b5
                    '''
            elif event == 2:
                sql = f'''
                    SELECT d.id, b5 as ptoken{column_str} FROM telem.i_search_damson d
                    left join telem.i_search_abiu a on d.d2=a.a8
                    left join telem.i_search_babaco b on a.a8=b.b4
                    {select_str}
                    where b4 is not null and b7=1{sql_mid}
                    group by b5
                    '''
            #! 測試全資料
            # sql = '''SELECT a.id, b5 as ptoken FROM telem.i_search_abiu a
            #         left join telem.i_search_babaco b on a.a8=b.b4
            #         where b.b4 is not null
            #         group by b4 limit 1000000;'''
            # print(sql)
            # return result
            # print(sql)
            qs = Babaco.objects.raw(sql)
            data_list = []
            data = {}
            count = 0
            for item in qs:
                #! 刪除已打過得電話
                if del_call_phone:
                    user_id = item.user_id if item.user_id else ''
                    if user_list and user_id in user_list:
                        continue
                phone = aes_decrypt(item.ptoken)
                phone = phone.replace('#', '')
                if len(phone) == 10:
                    count += 1
                    data_list.append({
                        'phone': phone,
                        })
            data['phone'] = {
                            'headers': [['phone']],
                            'rows': data_list,
                            }
            total_str = f'(yes){self.user.username}{sql}(home)'
            hl = hashlib.md5()
            hl.update(total_str.encode(encoding='utf-8'))
            filename = hl.hexdigest() + '.xlsx'
            msg = excel_file_write_sheets(filename, data)
            # if os.path.exists('output_file/' + filename):
            #     logger.info('檔案已存在')
            # else:
            #     msg = excel_file_write_sheets(filename, data)
            # os.remove('output_file/' + filename)

            #! 儲存歷史紀錄
            try:
                history = DownloadHistory.objects.get(filename=filename)
                downloads = history.downloads
                result['downloads'] = downloads + 1
                history.count = count
                history.downloads = downloads + 1
                history.last_download_time = datetime
                history.save()
            except:
                kwargs = {
                        'company': company,
                        'user': self.user,
                        'filename': filename,
                        'event': event,
                        'lbtype': lbtype,
                        'age_range': json.loads(age_interval) if age_interval else [],
                        'uid_tag': json.loads(gender) if gender else [],
                        'case_type': json.loads(credits) if credits else [],
                        'addr': json.loads(u_location) if u_location else [],
                        'lb_addr': json.loads(location) if location else [],
                        'setting_time': setting_time if setting_time else '',
                        'right_type': json.loads(rights) if rights and rights != '[]' else json.loads(right_type) if right_type else [],
                        'right_holder': right_holder if right_holder else '',
                        'set_amount_lo': set_amount_lo,
                        'set_amount_up': set_amount_up,
                        'set_start_time': {'set_start_time_lo': set_start_time_lo, 'set_start_time_up': set_start_time_up} if set_start_time_lo or set_start_time_up else {},
                        'set_end_time': {'set_end_time_lo': set_end_time_lo, 'set_end_time_up': set_end_time_up} if set_end_time_lo or set_end_time_up else {},
                        'count': count,
                        'last_download_time': datetime,
                        'del_call_phone': del_call_phone,
                }
                history = DownloadHistory.objects.create(**kwargs)
                result['downloads'] = 1

            result['filename'] = filename
            result['status'] = 'OK'
            result['msg'] = '檔案產製完成'
        except Exception as e:
            print(e, 'exception in line', sys.exc_info()[2].tb_lineno)
            result['msg'] = '發生不可預期的錯誤，請聯繫官方客服'
        return result

    @extend_schema(
        summary='清單下載',
        description='清單下載',
        request=DownloadPhoneNumberSerializer,
        responses={
            200: OpenApiResponse(description='ok'),
            401: OpenApiResponse(description='身分認證失敗'),
        },
    )

    def post(self, request):
        self.user = User.objects.get(username=request.user.get_username())
        logger.info('清單下載')
        time_start = time.perf_counter()
        result = self.process(request)
        time_end = time.perf_counter()
        logger.info(f'花費時間：{time_end - time_start}秒')
        if result['status'] == 'NG':
            return HttpResponseBadRequest(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")
        return HttpResponse(json.dumps(result, ensure_ascii=False, cls=CustomJsonEncoder), content_type="application/json; charset=utf-8")