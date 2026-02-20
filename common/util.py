import ast
import base64
import copy
import json
import logging
import os
import re
import time
import traceback
from datetime import date, datetime
from decimal import Decimal
from itertools import islice

import certifi
import openpyxl
import requests
import urllib3
from Crypto.Cipher import AES
from django.conf import settings
from django.db import close_old_connections, connections
from django.db.models.fields.files import ImageFieldFile
from openpyxl.utils import get_column_letter
from wsos_common.enums import LB_TYPE
from wsos_common.utility import getLBType

from users.models import Company, CompanyUserMapping, DownloadHistory, User

logger = logging.getLogger(__name__)
file_path = "output_file"
cities_str = 'ABCDEFGHIJKMNOPQTUVWXZ'
#! 測試用
test_sub_domain = 'www-test'
# test_sub_domain = 'zsa'

# batch for list
def batch(iterable, n=1):
    l = len(iterable)
    for ndx in range(0, l, n):
        yield iterable[ndx:min(ndx + n, l)]

# batch for dictionary
def chunks(iterable, n=1):
    it = iter(iterable)
    for i in range(0, len(iterable), n):
        yield {k:iterable[k] for k in islice(it, n)}

def set_column_widths(ws):
    column_widths = []
    for row_no, row in enumerate(ws.rows):
        for col_no, cell in enumerate(row):
            re_results = re.findall(r"\w|\s", str(cell.value), re.ASCII)
            cell_length = int( len(re_results) + ( len(str(cell.value)) - len(re_results) ) * 2 + 2 )
            if cell_length > 50:
                cell_length = 50
            if len(column_widths) > col_no:
                if cell_length > column_widths[col_no]:
                    column_widths[col_no] = cell_length
            else:
                column_widths.append(cell_length)
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width
    return ws

def excel_file_write_sheets(excel_filename, sheets_data):
    if not os.path.exists(file_path):
        os.mkdir(file_path)
    excel_filepath = os.path.join(file_path, excel_filename)
    try:
        wb = openpyxl.Workbook()
        sheets_names = list(sheets_data.keys())
        for sheet_title, sheet_data in sheets_data.items():
            # sheetnames = wb.sheetnames
            # ws = wb.worksheets[0]
            sheet_no = len(wb.worksheets)
            sheet_index = sheets_names.index(sheet_title)
            # print(sheet_no, sheet_index, sheet_title)
            if sheet_no < sheet_index + 1:
                ws = wb.create_sheet(sheet_title)
            else:
                ws = wb.active
                ws.title = sheet_title
            headers = sheet_data['headers']
            rows = sheet_data['rows']
            if rows == None or headers == None or len(rows) == 0:
                print('No rows or header to write to excel file!'.format())
            for header in headers:
                ws.append(header)
            for row in rows:
                if isinstance(row, dict):
                    ws.append(list(row.values()))
                else:
                    ws.append(list(row))
            ws = set_column_widths(ws)
            if 'freeze_panes' in sheet_data.keys():
                ws.freeze_panes = sheet_data['freeze_panes']
            if 'activeCell' in sheet_data.keys():
                ws.sheet_view.selection[0].activeCell = sheet_data['activeCell']
            ws = None
        wb.save(excel_filepath)
        # print ('{} excel file saved'.format(excel_filepath))
        return excel_filepath
    except IOError:
        return "{} excel file output error: {}".format(excel_filepath, traceback.print_exc())

#! 取得全台代碼(0：代碼找名稱，1：名稱找代碼)
def get_all_code(type=0):
    headers = {'Authorization': 'token fde85ab9e07f128e90d63091aca86f1e9a9d139d'}
    url = f'https://lbor.wsos.com.tw/common/car/get_all_code/?select_type={type}'
    code_datas = requests.get(url, headers=headers)
    code_datas = code_datas.json()
    return code_datas

#! 加密
def aes_encrypt(text):
    try:
        #! 金鑰
        key = settings.THUNDERBOLT_KEY
        key = key.encode('utf8')

        text = text.replace('\u3000', '').encode()
        cryptor = AES.new(key, AES.MODE_ECB)
        length = 16
        count = len(text)
        add = length - (count % length) if (count % length != 0) else 0
        text += '\0'.encode() * add
        cipher_text = cryptor.encrypt(text)
        encodestrs = base64.b64encode(cipher_text)
        enctext = encodestrs.decode('utf8')
    except:
        enctext = None
    return enctext

#! 解密
def aes_decrypt(text):
    try:
        #! 金鑰
        key = settings.THUNDERBOLT_KEY
        key = key.encode('utf8')

        cryptor = AES.new(key, AES.MODE_ECB)
        text = base64.b64decode(text)
        plain_text = cryptor.decrypt(text).decode()
        plain_text = plain_text.rstrip('\0')
    except:
        plain_text = None
    return plain_text

def check_role(request):
    #! 檢查sub_domain
    user_id = User.objects.get(username=request.user.get_username()).id
    urls = request.build_absolute_uri('/')[:-1].strip("/")
    if '.telem.com.tw' in urls or '.vvips.com.tw' in urls:
        url_t = '.telem.com.tw' if '.telem.com.tw' in urls else '.vvips.com.tw' if '.vvips.com.tw' in urls else ''
        sub_domain = urls.split(url_t)[0].split('//')[1]
    else:
        sub_domain = test_sub_domain
    companys = Company.objects.filter(sub_domain=sub_domain, is_valid=1)
    if companys:
        company_id = companys[0].id
    else:
        company_id = None
    # else:
    #     company_id = 1
    #! 檢查角色
    #* role_dict = {0: 'Administrator(元宏本身)', 1: Admin(老闆), 2: Manager(店東或廠商主管), 3: Operator(廠商旗下的業務), 4: other(非此sub_domain下的帳號)}
    check_user = CompanyUserMapping.objects.filter(user_id=user_id, is_valid=1)
    if check_user:
        role = 4 if company_id and check_user[0].company_id != company_id else 1 if check_user[0].is_admin == 1 else 2 if check_user[0].is_manager == 1 else 3
    else:
        role = 0
    return role

def get_sub_domain(request):
    urls = request.build_absolute_uri('/')[:-1].strip("/")
    if '.telem.com.tw' in urls or '.vvips.com.tw' in urls:
        url_t = '.telem.com.tw' if '.telem.com.tw' in urls else '.vvips.com.tw' if '.vvips.com.tw' in urls else ''
        sub_domain = urls.split(url_t)[0].split('//')[1]
    else:
        sub_domain = test_sub_domain
    return sub_domain

def pdata_handle(pdata):
    for pattern in ["^09\d{8}[123456789].*?$", "^09\d{8}[0][012345678].*?$", "^09\d{8}[0]$"]:
        matches = pdata if re.search(pattern, pdata) else ''
        if matches:
            break
    pdata_list = re.findall(r'09\d{8}', pdata) if not matches and re.findall(r'09\d{8}', pdata) else []
    return pdata_list
